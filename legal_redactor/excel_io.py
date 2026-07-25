from __future__ import annotations

from collections.abc import Callable
from io import BytesIO
from pathlib import PurePosixPath
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils.exceptions import InvalidFileException

from .models import MappingEntry, RedactionMap


EXCEL_INPUT_SUFFIXES = frozenset({".xlsx", ".xlsm"})
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExcelFormulaLeakError(ValueError):
    def __init__(self, locations: tuple[str, ...]) -> None:
        self.locations = locations
        super().__init__("Excel 公式包含待替换内容，已阻止源格式输出: " + ", ".join(locations))


def _safe_basename(filename: str) -> str:
    return PurePosixPath(str(filename).replace("\\", "/")).name or "上传文件"


def _load_workbook(content: bytes, filename: str):
    try:
        suffix = "." + _safe_basename(filename).rsplit(".", 1)[-1].lower() if "." in _safe_basename(filename) else ""
        if suffix not in EXCEL_INPUT_SUFFIXES:
            raise ValueError
        return load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=False,
            keep_links=False,
            keep_vba=False,
            rich_text=True,
        )
    except (BadZipFile, InvalidFileException, KeyError, ValueError) as exc:
        raise ValueError(
            f"读取文件 {_safe_basename(filename)} 失败: Excel 格式无效、已加密或文件已损坏"
        ) from exc


def _cell_text(value: object) -> str | None:
    if isinstance(value, CellRichText):
        return str(value)
    if isinstance(value, str):
        return value
    return None


def extract_workbook_text(content: bytes, filename: str) -> str:
    workbook = _load_workbook(content, filename)
    lines: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    continue
                text = _cell_text(cell.value)
                if text:
                    lines.append(text)
    return "\n".join(lines)


def _formula_locations(workbook, redaction_map: RedactionMap) -> tuple[str, ...]:
    originals = tuple(entry.original for entry in redaction_map.mappings if entry.original)
    if not originals:
        return ()
    locations: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str):
                    if any(original in cell.value for original in originals):
                        locations.append(f"{worksheet.title}!{cell.coordinate}")
    return tuple(locations)


def redact_workbook(
    content: bytes,
    filename: str,
    redaction_map: RedactionMap,
    apply_mappings: Callable[[str, list[MappingEntry]], str],
) -> bytes:
    workbook = _load_workbook(content, filename)
    locations = _formula_locations(workbook, redaction_map)
    if locations:
        raise ExcelFormulaLeakError(locations)

    mappings = list(redaction_map.mappings)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    continue
                text = _cell_text(cell.value)
                if text is None:
                    continue
                replaced = apply_mappings(text, mappings)
                if replaced != text:
                    cell.value = replaced

    calculation = workbook.calculation
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True
    calculation.calcCompleted = False
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

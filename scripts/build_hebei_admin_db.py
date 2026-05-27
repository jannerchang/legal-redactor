#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import sqlite3
import ssl
import urllib.request
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DIVISIONS_CSV = ROOT / "data" / "hebei_admin_divisions.csv"
DEFAULT_ALIASES_CSV = ROOT / "data" / "hebei_aliases.csv"
DEFAULT_DB = ROOT / "data" / "hebei_admin_divisions.sqlite"
XZQH_INDEX_2023 = "https://www.xzqh.org/show/china/2023/13/index.html"

DIVISION_FIELDS = [
    "code",
    "name",
    "full_name",
    "level",
    "parent_code",
    "parent_name",
    "city_name",
    "county_name",
    "township_name",
    "village_name",
    "entity_type",
    "urban_rural_code",
    "source",
    "source_year",
    "is_active",
]

ALIAS_FIELDS = [
    "alias",
    "canonical_name",
    "division_code",
    "alias_type",
    "source",
    "confidence",
]


class StatsTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_td = False
        self.current_row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self.current_row = []
        if tag.lower() == "td":
            self.in_td = True

    def handle_data(self, data: str) -> None:
        if self.in_td:
            value = data.strip()
            if value:
                self.current_row.append(value)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "td":
            self.in_td = False
        if tag.lower() == "tr" and self.current_row:
            self.rows.append(self.current_row)


class LinkAndTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self.text_parts: list[str] = []
        self._href: str | None = None
        self._link_text: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style"}:
            self._skip_depth += 1
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._link_text = []

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        value = data.strip()
        if not value:
            return
        self.text_parts.append(value)
        if self._href is not None:
            self._link_text.append(value)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style"} and self._skip_depth:
            self._skip_depth -= 1
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, "".join(self._link_text).strip()))
            self._href = None
            self._link_text = []


def main() -> int:
    parser = argparse.ArgumentParser(description="构建河北省行政区划/基层组织 SQLite 数据库")
    parser.add_argument("--divisions-csv", default=str(DEFAULT_DIVISIONS_CSV), help="本地行政区划 CSV")
    parser.add_argument("--aliases-csv", default=str(DEFAULT_ALIASES_CSV), help="本地别名 CSV")
    parser.add_argument("--db", default=str(DEFAULT_DB), help="输出 SQLite 路径")
    parser.add_argument("--html", action="append", default=[], help="本地国家统计局 HTML 文件，可重复传入")
    parser.add_argument("--url", action="append", default=[], help="国家统计局 HTML URL，可重复传入")
    parser.add_argument("--excel", action="append", default=[], help="本地 Excel 文件；需要 openpyxl，可重复传入")
    parser.add_argument("--fetch-xzqh-2023", action="store_true", help="从区划地名网批量抓取 2023 河北县级页面作为统计区划转载数据源")
    parser.add_argument("--stats", action="store_true", help="输出当前 SQLite 统计信息后退出")
    args = parser.parse_args()

    if args.stats:
        _print_stats(Path(args.db))
        return 0

    divisions = list(_read_divisions_csv(Path(args.divisions_csv)))
    aliases = list(_read_aliases_csv(Path(args.aliases_csv)))
    divisions.extend(_read_html_sources(args.html, args.url))
    divisions.extend(_read_excel_sources(args.excel))
    if args.fetch_xzqh_2023:
        divisions.extend(_fetch_xzqh_2023())
    divisions = _dedupe_divisions(divisions)
    aliases.extend(_generate_aliases(divisions))
    aliases = _dedupe_aliases(aliases)
    _write_sqlite(Path(args.db), divisions, aliases)
    _write_csv(Path(args.divisions_csv), DIVISION_FIELDS, divisions)
    _write_csv(Path(args.aliases_csv), ALIAS_FIELDS, aliases)
    print(f"wrote: {args.db}")
    print(f"divisions: {len(divisions)}")
    print(f"aliases: {len(aliases)}")
    _print_stats(Path(args.db))
    return 0


def _read_divisions_csv(path: Path) -> Iterable[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_aliases_csv(path: Path) -> Iterable[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_html_sources(paths: list[str], urls: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in paths:
        rows.extend(_parse_stats_html(Path(path).read_text(encoding="utf-8", errors="ignore"), "local_html"))
    for url in urls:
        try:
            request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(request, timeout=30, context=ssl._create_unverified_context()) as response:
                html = response.read().decode("utf-8", errors="ignore")
        except OSError as exc:
            print(f"warning: failed to fetch {url}: {exc}")
            continue
        rows.extend(_parse_stats_html(html, url))
    return rows


def _fetch_xzqh_2023() -> list[dict[str, str]]:
    index_html = _fetch_text(XZQH_INDEX_2023)
    parser = LinkAndTextParser()
    parser.feed(index_html)
    urls: dict[str, str] = {}
    city_rows: dict[str, dict[str, str]] = {}
    for href, name in parser.links:
        match = re.search(r"(13\d{2,4})\.html$", href)
        if not match:
            continue
        code = match.group(1)
        url = urljoin(XZQH_INDEX_2023, href)
        if len(code) == 4 or (len(code) == 6 and code.endswith("00")):
            city_code = code[:4] + "00000000"
            city_rows[city_code] = _division_row(
                code=city_code,
                name=name,
                level="city",
                parent_code="130000000000",
                parent_name="河北省",
                city_name=name,
                source=url,
                source_year="2023",
            )
        elif len(code) == 6:
            urls[code] = url

    rows: list[dict[str, str]] = [
        _division_row(
            code="130000000000",
            name="河北省",
            level="province",
            source=XZQH_INDEX_2023,
            source_year="2023",
        )
    ]
    rows.extend(city_rows.values())

    for county_code, url in sorted(urls.items()):
        try:
            rows.extend(_parse_xzqh_county_page(_fetch_text(url), url, city_rows))
        except Exception as exc:  # noqa: BLE001 - one county page must not abort the build.
            print(f"warning: failed to parse {url}: {exc}")
    return rows


def _fetch_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=30, context=ssl._create_unverified_context()) as response:
        raw = response.read()
    for encoding in ("utf-8", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _parse_xzqh_county_page(html: str, source: str, city_rows: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    parser = LinkAndTextParser()
    parser.feed(html)
    text = "\n".join(parser.text_parts)
    raw_items = _parse_code_name_lines(text, source)
    if not raw_items:
        return []

    rows: dict[str, dict[str, str]] = {}
    for item in raw_items:
        code = item["code"]
        level = _level_from_code(code, item["name"])
        if level == "province":
            continue
        city_code = code[:4] + "00000000"
        city_name = city_rows.get(city_code, {}).get("name", "")
        county_code = code[:6] + "000000"
        township_code = code[:9] + "000"
        parent_code = _parent_code_for(code, level)
        parent_name = rows.get(parent_code, {}).get("name") or city_rows.get(parent_code, {}).get("name", "")
        if level == "county" and not parent_name:
            parent_name = city_name
        county_name = item["name"] if level in {"county", "county_city"} else rows.get(county_code, {}).get("name", "")
        township_name = item["name"] if level == "township" else rows.get(township_code, {}).get("name", "")
        village_name = item["name"] if level in {"village", "community"} else ""
        full_name = _compose_full_name(level, item["name"], city_name, county_name, township_name)
        rows[code] = _division_row(
            code=code,
            name=item["name"],
            level=level,
            parent_code=parent_code,
            parent_name=parent_name,
            city_name=city_name,
            county_name=county_name,
            township_name=township_name,
            village_name=village_name,
            urban_rural_code=item.get("urban_rural_code", ""),
            source=source,
            source_year="2023",
            full_name=full_name,
            entity_type=_entity_type(level, item["name"]),
        )
    return list(rows.values())


def _parse_code_name_lines(text: str, source: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    pattern = re.compile(r"(?m)^\s*(13\d{10})\s+(?:(\d{3})\s+)?([^\n\r]+?)\s*$")
    for match in pattern.finditer(text):
        name = _clean_imported_name(match.group(3))
        if not name or len(name) > 40:
            continue
        rows.append(
            {
                "code": match.group(1),
                "urban_rural_code": match.group(2) or "",
                "name": name,
                "source": source,
            }
        )
    return rows


def _clean_imported_name(value: str) -> str:
    value = re.sub(r"\[[^\]]+\]", "", value)
    value = re.sub(r"（.*?）|\(.*?\)", "", value)
    value = value.strip(" ：:，,。；;　\t")
    return value


def _parse_stats_html(html: str, source: str) -> list[dict[str, str]]:
    parser = StatsTableParser()
    parser.feed(html)
    rows: list[dict[str, str]] = []
    for cells in parser.rows:
        if len(cells) < 2:
            continue
        code = cells[0]
        name = cells[-1]
        if not code.isdigit() or not name:
            continue
        level = _level_from_code(code, name)
        rows.append(
            {
                "code": code,
                "name": name,
                "full_name": name,
                "level": level,
                "parent_code": "",
                "parent_name": "",
                "city_name": "",
                "county_name": "",
                "township_name": "",
                "village_name": name if level in {"village", "community"} else "",
                "entity_type": _entity_type(level, name),
                "urban_rural_code": cells[1] if len(cells) > 2 and cells[1].isdigit() else "",
                "source": source,
                "source_year": "",
                "is_active": "1",
            }
        )
    return rows


def _read_excel_sources(paths: list[str]) -> list[dict[str, str]]:
    if not paths:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("warning: openpyxl not installed; skipped Excel imports")
        return []
    rows: list[dict[str, str]] = []
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        header = [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True))]
        for raw in sheet.iter_rows(values_only=True):
            item = {header[index]: str(value or "").strip() for index, value in enumerate(raw) if index < len(header)}
            if item.get("code") and item.get("name"):
                rows.append({field: item.get(field, "") for field in DIVISION_FIELDS})
    return rows


def _dedupe_divisions(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: dict[str, dict[str, str]] = {}
    for row in rows:
        code = row.get("code", "")
        if not code:
            continue
        previous = deduped.get(code)
        if previous is None or _source_rank(row.get("source", "")) >= _source_rank(previous.get("source", "")):
            deduped[code] = {field: row.get(field, "") for field in DIVISION_FIELDS}
    return list(deduped.values())


def _dedupe_aliases(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows:
        key = (row.get("alias", ""), row.get("division_code", ""))
        if not key[0] or not key[1]:
            continue
        previous = deduped.get(key)
        if previous is None or float(row.get("confidence") or 0) >= float(previous.get("confidence") or 0):
            deduped[key] = {field: row.get(field, "") for field in ALIAS_FIELDS}
    return list(deduped.values())


def _source_rank(source: str) -> int:
    if "xzqh.org" in source:
        return 3
    if source and source != "seed":
        return 2
    return 1


def _division_row(
    code: str,
    name: str,
    level: str,
    parent_code: str = "",
    parent_name: str = "",
    city_name: str = "",
    county_name: str = "",
    township_name: str = "",
    village_name: str = "",
    entity_type: str | None = None,
    urban_rural_code: str = "",
    source: str = "",
    source_year: str = "",
    full_name: str | None = None,
) -> dict[str, str]:
    return {
        "code": code,
        "name": name,
        "full_name": full_name or _compose_full_name(level, name, city_name, county_name, township_name),
        "level": level,
        "parent_code": parent_code,
        "parent_name": parent_name,
        "city_name": city_name,
        "county_name": county_name,
        "township_name": township_name,
        "village_name": village_name,
        "entity_type": entity_type or _entity_type(level, name),
        "urban_rural_code": urban_rural_code,
        "source": source,
        "source_year": source_year,
        "is_active": "1",
    }


def _parent_code_for(code: str, level: str) -> str:
    if level == "city":
        return "130000000000"
    if level in {"county", "county_city"}:
        return code[:4] + "00000000"
    if level == "township":
        return code[:6] + "000000"
    if level in {"village", "community"}:
        return code[:9] + "000"
    return ""


def _compose_full_name(level: str, name: str, city_name: str = "", county_name: str = "", township_name: str = "") -> str:
    if level == "province":
        return name
    if level == "city":
        return "河北省" + name if not name.startswith("河北省") else name
    if level in {"county", "county_city"}:
        return "河北省" + city_name + name
    if level == "township":
        return "河北省" + city_name + county_name + name
    if level in {"village", "community"}:
        return "河北省" + city_name + county_name + township_name + _grassroots_full_name(name, level)
    return "河北省" + city_name + county_name + township_name + name


def _grassroots_full_name(name: str, level: str) -> str:
    if name.endswith(("村民委员会", "居民委员会", "村委会", "居委会")):
        return name.replace("村委会", "村民委员会").replace("居委会", "居民委员会")
    if name.endswith("社区") or level == "community":
        return name + "居民委员会"
    if name.endswith("村") or level == "village":
        return name + "村民委员会"
    return name


def _generate_aliases(divisions: list[dict[str, str]]) -> list[dict[str, str]]:
    aliases: dict[tuple[str, str], dict[str, str]] = {}
    for row in divisions:
        code = row.get("code", "")
        full_name = row.get("full_name") or row.get("name", "")
        name = row.get("name", "")
        for alias, alias_type, confidence in _aliases_for_name(name, full_name):
            if not alias or alias == full_name:
                continue
            aliases[(alias, code)] = {
                "alias": alias,
                "canonical_name": full_name,
                "division_code": code,
                "alias_type": alias_type,
                "source": "generated",
                "confidence": str(confidence),
            }
    return list(aliases.values())


def _aliases_for_name(name: str, full_name: str) -> Iterable[tuple[str, str, float]]:
    yield name, "short_name", 0.9
    if full_name.startswith("河北省"):
        yield full_name.removeprefix("河北省"), "without_province", 0.96
    parts = _strip_hebei_prefix_combinations(full_name)
    for alias, confidence in parts:
        yield alias, "continuous_combo", confidence
    if "居民委员会" in name:
        yield name.replace("居民委员会", "居委会"), "org_variant", 0.94
    if "村民委员会" in name:
        yield name.replace("村民委员会", "村委会"), "org_variant", 0.94
        yield name.replace("村民委员会", "村"), "short_local_name", 0.82
    if name.endswith("社区"):
        yield name + "居民委员会", "org_full_variant", 0.94
        yield name + "居委会", "org_variant", 0.9
    if name.endswith("村"):
        yield name + "村民委员会", "org_full_variant", 0.94
        yield name + "村委会", "org_variant", 0.9
    if "街道办事处" in name:
        yield name.replace("街道办事处", "街道"), "org_variant", 0.94
    if "社区居民委员会" in name:
        yield name.replace("社区居民委员会", "社区"), "short_local_name", 0.86


def _strip_hebei_prefix_combinations(full_name: str) -> Iterable[tuple[str, float]]:
    value = full_name.removeprefix("河北省")
    if value != full_name:
        yield value, 0.96
    match = re.match(r"[^市]+市(.+)", value)
    if match:
        yield match.group(1), 0.92
    match = re.match(r"[^市]+市(?:[^区县市]+[区县市]|[^自治县]+自治县)(.+)", value)
    if match:
        yield match.group(1), 0.9


def _write_sqlite(path: Path, divisions: list[dict[str, str]], aliases: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as conn:
        conn.execute("DROP TABLE IF EXISTS admin_divisions")
        conn.execute("DROP TABLE IF EXISTS admin_aliases")
        conn.execute(
            """
            CREATE TABLE admin_divisions (
                code TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                full_name TEXT NOT NULL,
                level TEXT NOT NULL,
                parent_code TEXT,
                parent_name TEXT,
                city_name TEXT,
                county_name TEXT,
                township_name TEXT,
                village_name TEXT,
                entity_type TEXT NOT NULL,
                urban_rural_code TEXT,
                source TEXT,
                source_year TEXT,
                is_active INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE admin_aliases (
                alias TEXT NOT NULL,
                canonical_name TEXT NOT NULL,
                division_code TEXT NOT NULL,
                alias_type TEXT,
                source TEXT,
                confidence REAL,
                UNIQUE(alias, division_code)
            )
            """
        )
        conn.executemany(
            f"INSERT OR REPLACE INTO admin_divisions ({','.join(DIVISION_FIELDS)}) VALUES ({','.join('?' for _ in DIVISION_FIELDS)})",
            [[row.get(field, "") for field in DIVISION_FIELDS] for row in divisions if row.get("code")],
        )
        conn.executemany(
            f"INSERT OR REPLACE INTO admin_aliases ({','.join(ALIAS_FIELDS)}) VALUES ({','.join('?' for _ in ALIAS_FIELDS)})",
            [[row.get(field, "") for field in ALIAS_FIELDS] for row in aliases if row.get("alias")],
        )
        conn.execute("CREATE INDEX idx_admin_divisions_name ON admin_divisions(name)")
        conn.execute("CREATE INDEX idx_admin_aliases_alias ON admin_aliases(alias)")


def _write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def _level_from_code(code: str, name: str) -> str:
    if code == "130000000000":
        return "province"
    if name.endswith(("村民委员会", "居民委员会")):
        return "village" if name.endswith("村民委员会") else "community"
    if not code.endswith("000") and len(code) == 12:
        if name.endswith(("社区", "居委会", "居民委员会")):
            return "community"
        return "village"
    if code.endswith("00000000") and not code.endswith("0000000000"):
        return "city"
    if code.endswith("000000"):
        return "county_city" if name.endswith("市") else "county"
    if code.endswith("0000000000"):
        return "city"
    if code.endswith("000"):
        return "township"
    return "village"


def _entity_type(level: str, name: str) -> str:
    if level in {"village", "community"} or name.endswith(("村民委员会", "居民委员会")):
        return "grassroots_org"
    return "location"


def _print_stats(path: Path) -> None:
    if not path.exists():
        print(f"database not found: {path}")
        return
    with sqlite3.connect(path) as conn:
        city_count = _count(conn, "level = 'city'")
        county_count = _count(conn, "level IN ('county', 'county_city')")
        township_count = _count(conn, "level = 'township'")
        village_community_count = _count(conn, "level IN ('village', 'community')")
        village_committee_count = _count(conn, "level = 'village'")
        resident_committee_count = _count(conn, "level = 'community'")
        alias_count = conn.execute("SELECT COUNT(*) FROM admin_aliases").fetchone()[0]
    print("stats:")
    print(f"  市数量: {city_count}")
    print(f"  区县数量: {county_count}")
    print(f"  乡镇街道数量: {township_count}")
    print(f"  村 / 社区数量: {village_community_count}")
    print(f"  村民委员会数量: {village_committee_count}")
    print(f"  居民委员会数量: {resident_committee_count}")
    print(f"  alias 数量: {alias_count}")


def _count(conn: sqlite3.Connection, where: str) -> int:
    return int(conn.execute(f"SELECT COUNT(*) FROM admin_divisions WHERE {where} AND is_active = 1").fetchone()[0])


if __name__ == "__main__":
    raise SystemExit(main())
